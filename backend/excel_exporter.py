"""
Exporteur Excel — génère les fichiers AMDEC et Gamme à partir des données JSON.

Utilise le template Templates_AMDEC_Gamme.xlsx comme base de mise en forme.
Copie la structure + styles du template, puis injecte les données générées.
"""

import logging
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
TEMPLATE_PATH = BASE_DIR / "templates" / "Templates_AMDEC_Gamme.xlsx"
EXPORTS_DIR = BASE_DIR / "exports"

logger = logging.getLogger(__name__)


def _ouvrir_template_ou_creer(
    chemin_sortie: Path, feuille_cible: str
) -> tuple:
    """
    Copie le template Excel dans chemin_sortie et ouvre la feuille `feuille_cible`.
    Si le template est absent, crée un workbook vide avec la feuille demandée.
    Retourne (workbook, worksheet).
    """
    if TEMPLATE_PATH.exists():
        shutil.copy2(TEMPLATE_PATH, chemin_sortie)
        wb = openpyxl.load_workbook(chemin_sortie)
        if feuille_cible in wb.sheetnames:
            ws = wb[feuille_cible]
        else:
            logger.warning(
                "Feuille '%s' absente du template — création d'une feuille vide.", feuille_cible
            )
            ws = wb.create_sheet(feuille_cible)
    else:
        logger.warning(
            "Template introuvable (%s) — export sans mise en forme du template.", TEMPLATE_PATH
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = feuille_cible
    return wb, ws

# Couleurs
ROUGE = "C00000"
ORANGE = "FF6600"
VERT = "00B050"
BLEU_HEADER = "1F4E79"
GRIS_CLAIR = "F2F2F2"


def _ipr_couleur(ipr: int) -> str:
    if ipr > 100:
        return ROUGE
    if ipr > 40:
        return ORANGE
    return VERT


def _ecrire_cellule(ws, row, col, valeur, bold=False, centre=False, fill_hex=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=valeur)
    if bold:
        cell.font = Font(bold=True)
    if centre:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    elif wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    return cell


def _ecrire_voisin_droit(ws, label_cell, valeur, max_offset: int = 5):
    """
    Écrit `valeur` dans la cellule libre la plus proche à droite de label_cell.
    Saute les MergedCell (lecture seule) et trouve la cellule de tête de zone fusionnée.
    """
    from openpyxl.cell.cell import MergedCell

    for offset in range(1, max_offset + 1):
        target = ws.cell(row=label_cell.row, column=label_cell.column + offset)
        if isinstance(target, MergedCell):
            # Trouver la cellule maître de la zone fusionnée qui contient target
            for merged_range in ws.merged_cells.ranges:
                if (target.coordinate in merged_range):
                    master = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if master.coordinate != label_cell.coordinate:
                        master.value = valeur
                        return
                    break
            continue
        target.value = valeur
        return


def _remplir_entete(ws, designation: str, reference: str, traitements: str):
    """Remplit le bloc d'identification en haut de chaque feuille."""
    today = date.today().strftime("%d/%m/%Y")

    mapping = {
        "Désignation :": designation,
        "Référence produit :": reference,
        "Traitement(s) :": traitements,
        "Date création :": today,
        "Statut :": "À valider",
    }

    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            for label, valeur in mapping.items():
                if label in val:
                    try:
                        _ecrire_voisin_droit(ws, cell, valeur)
                    except (AttributeError, ValueError):
                        pass


def exporter_amdec_produit(
    donnees: dict,
    nom_fichier: str,
    exports_dir: Path = None,
) -> Path:
    """
    Génère le fichier Excel AMDEC Produit à partir des données JSON.
    Retourne le chemin du fichier créé.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)

    chemin_sortie = exports_dir / nom_fichier
    wb, ws = _ouvrir_template_ou_creer(chemin_sortie, "AMDEC Produit")

    traitements_str = ", ".join(donnees.get("traitements_source", ["CN", "AR double face"]))
    _remplir_entete(ws, donnees.get("designation", ""), donnees.get("reference", ""), traitements_str)

    # Ligne de départ des données (après les en-têtes du template)
    ROW_START = 10
    modes = donnees.get("modes_defaillance", [])

    for i, mode in enumerate(modes):
        row = ROW_START + i

        g = int(mode.get("G", 1))
        o = int(mode.get("O", 1))
        d = int(mode.get("D", 1))
        ipr = g * o * d

        valeurs = [
            mode.get("no", str(i + 1).zfill(2)),
            mode.get("fonction_exigence", ""),
            mode.get("caracteristique_critique", ""),
            mode.get("mode_defaillance", ""),
            mode.get("effets_defaillance", ""),
            g,
            mode.get("causes_defaillance", ""),
            o,
            mode.get("controles_existants", ""),
            d,
            ipr,
            mode.get("classe_criticite", ""),
            mode.get("actions_correctives", ""),
            mode.get("responsable", ""),
            mode.get("delai", ""),
            mode.get("actions_realisees", ""),
            None,  # G'
            None,  # O'
            None,  # D'
        ]

        for col, valeur in enumerate(valeurs, start=1):
            centre = col in (1, 6, 8, 10, 11)
            cell = _ecrire_cellule(ws, row, col, valeur, centre=centre, wrap=True)

            # Colorer la cellule IPR selon criticité
            if col == 11 and isinstance(valeur, int):
                cell.fill = PatternFill("solid", fgColor=_ipr_couleur(valeur))
                cell.font = Font(bold=True, color="FFFFFF")

    # Ajuster hauteur des lignes données
    for row in range(ROW_START, ROW_START + len(modes)):
        ws.row_dimensions[row].height = 45

    # Avertissements en bas si présents
    avertissements = donnees.get("avertissements_generateur", [])
    if avertissements:
        row_avert = ROW_START + len(modes) + 2
        ws.cell(row=row_avert, column=1).value = "⚠️ Avertissements génération IA :"
        ws.cell(row=row_avert, column=1).font = Font(bold=True, color=ORANGE)
        for j, avert in enumerate(avertissements):
            ws.cell(row=row_avert + 1 + j, column=1).value = f"• {avert}"

    # Supprimer les autres feuilles pour l'export dédié
    for sheet_name in wb.sheetnames:
        if sheet_name != "AMDEC Produit":
            del wb[sheet_name]

    wb.save(chemin_sortie)
    return chemin_sortie


def exporter_amdec_process(
    donnees: dict,
    nom_fichier: str,
    exports_dir: Path = None,
) -> Path:
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)

    chemin_sortie = exports_dir / nom_fichier
    wb, ws = _ouvrir_template_ou_creer(chemin_sortie, "AMDEC Process")

    traitements_str = ", ".join(donnees.get("traitements_source", ["CN", "AR double face"]))
    _remplir_entete(ws, donnees.get("designation", ""), donnees.get("reference", ""), traitements_str)

    ROW_START = 10
    modes = donnees.get("modes_defaillance", [])

    for i, mode in enumerate(modes):
        row = ROW_START + i

        g = int(mode.get("G", 1))
        o = int(mode.get("O", 1))
        d = int(mode.get("D", 1))
        ipr = g * o * d

        valeurs = [
            mode.get("no", str(i + 1).zfill(2)),
            mode.get("operation_process", ""),
            mode.get("etape_poste", ""),
            mode.get("mode_defaillance", ""),
            mode.get("effets_produit", ""),
            g,
            mode.get("causes_process", ""),
            o,
            mode.get("controles_process", ""),
            d,
            ipr,
            mode.get("classe_criticite", ""),
            mode.get("actions_correctives", ""),
            mode.get("responsable", ""),
            mode.get("delai", ""),
            mode.get("actions_realisees", ""),
            mode.get("parametre_cle", ""),
            mode.get("valeur_cible", ""),
            None,  # G'
            None,  # IPR'
        ]

        for col, valeur in enumerate(valeurs, start=1):
            centre = col in (1, 6, 8, 10, 11)
            cell = _ecrire_cellule(ws, row, col, valeur, centre=centre, wrap=True)

            if col == 11 and isinstance(valeur, int):
                cell.fill = PatternFill("solid", fgColor=_ipr_couleur(valeur))
                cell.font = Font(bold=True, color="FFFFFF")

    for row in range(ROW_START, ROW_START + len(modes)):
        ws.row_dimensions[row].height = 45

    avertissements = donnees.get("avertissements_generateur", [])
    if avertissements:
        row_avert = ROW_START + len(modes) + 2
        ws.cell(row=row_avert, column=1).value = "⚠️ Avertissements génération IA :"
        ws.cell(row=row_avert, column=1).font = Font(bold=True, color=ORANGE)
        for j, avert in enumerate(avertissements):
            ws.cell(row=row_avert + 1 + j, column=1).value = f"• {avert}"

    for sheet_name in wb.sheetnames:
        if sheet_name != "AMDEC Process":
            del wb[sheet_name]

    wb.save(chemin_sortie)
    return chemin_sortie


def exporter_gamme(
    donnees: dict,
    nom_fichier: str,
    exports_dir: Path = None,
) -> Path:
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)

    chemin_sortie = exports_dir / nom_fichier
    wb, ws = _ouvrir_template_ou_creer(chemin_sortie, "Gamme Production")

    traitements_str = ", ".join(donnees.get("traitements_source", ["CN", "AR double face"]))
    _remplir_entete(ws, donnees.get("designation", ""), donnees.get("reference", ""), traitements_str)

    ROW_START = 11  # La gamme commence ligne 11 dans le template
    operations = donnees.get("operations", [])

    for i, op in enumerate(operations):
        row = ROW_START + i

        valeurs = [
            op.get("no_op", ""),
            op.get("designation", ""),
            op.get("description_detaillee", ""),
            op.get("poste_machine", ""),
            op.get("outillage_fixture", ""),
            op.get("parametre_1", ""),
            op.get("valeur_1", ""),
            op.get("parametre_2", ""),
            op.get("valeur_2", ""),
            op.get("parametre_3", ""),
            op.get("temps_min", ""),
            op.get("point_controle", ""),
            op.get("moyen_controle", ""),
            op.get("frequence", ""),
            op.get("critere_acceptation", ""),
            op.get("ref_dit", ""),
            op.get("ref_infor", ""),
            None,  # Visa OP (manuel)
        ]

        for col, valeur in enumerate(valeurs, start=1):
            centre = col in (1, 11, 14, 16, 17)
            _ecrire_cellule(ws, row, col, valeur, centre=centre, wrap=True)

    for row in range(ROW_START, ROW_START + len(operations)):
        ws.row_dimensions[row].height = 50

    avertissements = donnees.get("avertissements_generateur", [])
    if avertissements:
        row_avert = ROW_START + len(operations) + 2
        ws.cell(row=row_avert, column=1).value = "⚠️ Avertissements génération IA :"
        ws.cell(row=row_avert, column=1).font = Font(bold=True, color=ORANGE)
        for j, avert in enumerate(avertissements):
            ws.cell(row=row_avert + 1 + j, column=1).value = f"• {avert}"

    for sheet_name in wb.sheetnames:
        if sheet_name != "Gamme Production":
            del wb[sheet_name]

    wb.save(chemin_sortie)
    return chemin_sortie


def exporter_plan_controle(
    donnees: dict,
    nom_fichier: str,
    exports_dir: Path = None,
) -> Path:
    """Exporte le Plan de Contrôle vers un fichier Excel."""
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)

    chemin_sortie = exports_dir / nom_fichier
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Contrôle"

    # ── En-tête ──────────────────────────────────────────────────────────────
    BLEU = "1B3A6B"
    BLANC = "FFFFFF"
    GRIS = "F2F2F2"
    ORANGE_HDR = "FF6600"

    entete_style = Font(bold=True, color=BLANC)
    entete_fill = PatternFill("solid", fgColor=BLEU)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=1, column=1, value="PLAN DE CONTRÔLE").font = Font(bold=True, size=14, color=BLEU)
    ws.cell(row=2, column=1, value=f"Référence : {donnees.get('reference', '')}")
    ws.cell(row=3, column=1, value=f"Désignation : {donnees.get('designation', '')}")
    ws.cell(row=4, column=1, value=f"Version : {donnees.get('version', 'A')} — Date : {donnees.get('date_creation', date.today().isoformat())}")

    # ── Colonnes ─────────────────────────────────────────────────────────────
    COLONNES = [
        ("N°", 5), ("Phase", 14), ("Opération Gamme", 20), ("Caractéristique", 20),
        ("Critère acceptation", 20), ("Moyen de contrôle", 18), ("Fréquence", 12),
        ("Responsable", 14), ("Enregistrement", 16), ("Action NC", 20),
    ]

    ROW_HDR = 6
    for col, (label, width) in enumerate(COLONNES, start=1):
        cell = ws.cell(row=ROW_HDR, column=col, value=label)
        cell.font = entete_style
        cell.fill = entete_fill
        cell.alignment = centre
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Données ───────────────────────────────────────────────────────────────
    points = donnees.get("points_controle", [])
    for i, pt in enumerate(points):
        row = ROW_HDR + 1 + i
        fill = PatternFill("solid", fgColor=GRIS) if i % 2 == 0 else None
        valeurs = [
            pt.get("id", i + 1),
            pt.get("phase", ""),
            pt.get("operation_gamme", ""),
            pt.get("caracteristique", ""),
            pt.get("critere_acceptation", ""),
            pt.get("moyen_controle", ""),
            pt.get("frequence", ""),
            pt.get("responsable", ""),
            pt.get("enregistrement", ""),
            pt.get("action_non_conformite", ""),
        ]
        for col, val in enumerate(valeurs, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
        ws.row_dimensions[row].height = 40

    # ── Avertissements ────────────────────────────────────────────────────────
    avertissements = donnees.get("avertissements_generateur", [])
    if avertissements:
        row_a = ROW_HDR + len(points) + 2
        ws.cell(row=row_a, column=1, value="⚠️ Avertissements génération IA :").font = Font(bold=True, color=ORANGE_HDR)
        for j, avert in enumerate(avertissements):
            ws.cell(row=row_a + 1 + j, column=1, value=f"• {avert}")

    wb.save(chemin_sortie)
    return chemin_sortie


def exporter_dossier_complet(dossier: dict, prefixe: str, exports_dir: Path = None) -> dict:
    """
    Exporte les 4 documents d'un dossier généré.
    Retourne un dict avec les chemins des fichiers créés.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR

    chemins = {}

    if "amdec_produit" in dossier:
        chemin = exporter_amdec_produit(
            dossier["amdec_produit"],
            f"{prefixe}_AMDEC_Produit.xlsx",
            exports_dir,
        )
        chemins["amdec_produit"] = chemin

    if "amdec_process" in dossier:
        chemin = exporter_amdec_process(
            dossier["amdec_process"],
            f"{prefixe}_AMDEC_Process.xlsx",
            exports_dir,
        )
        chemins["amdec_process"] = chemin

    if "gamme" in dossier:
        chemin = exporter_gamme(
            dossier["gamme"],
            f"{prefixe}_Gamme.xlsx",
            exports_dir,
        )
        chemins["gamme"] = chemin

    if "plan_controle" in dossier:
        chemin = exporter_plan_controle(
            dossier["plan_controle"],
            f"{prefixe}_Plan_Controle.xlsx",
            exports_dir,
        )
        chemins["plan_controle"] = chemin

    return chemins


def exporter_retouches(ref_article: str, nom_fichier: str = None, exports_dir: Path = None) -> Path:
    """Exporte les retouches d'un article vers un fichier Excel."""
    from backend.retouche_manager import (
        lister_retouches, stats_retouches, RESULTATS
    )

    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)

    if not nom_fichier:
        slug = ref_article.strip().replace("/", "_")
        nom_fichier = f"{slug}_Retouches.xlsx"

    chemin_sortie = exports_dir / nom_fichier
    retouches = lister_retouches(ref_article)
    stats = stats_retouches(ref_article)
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Liste des retouches ──────────────────────────────────────
    ws_liste = wb.active
    ws_liste.title = "Retouches"

    ws_liste["A1"] = f"FICHE RETOUCHES — {ref_article}"
    ws_liste["A1"].font = Font(bold=True, size=13)
    ws_liste["A2"] = f"Exporte le {date.today().strftime('%d/%m/%Y')} · {len(retouches)} retouche(s)"
    ws_liste["A2"].font = Font(italic=True, color="666666")
    ws_liste.row_dimensions[1].height = 22
    ws_liste.row_dimensions[2].height = 16

    entetes = [
        "ID Retouche", "Date", "Operateur", "Poste",
        "Defaut constate", "Operation retouche", "Resultat", "Commentaire"
    ]
    largeurs = [22, 14, 20, 12, 40, 40, 18, 35]
    FILL_HEADER = PatternFill("solid", fgColor="1F4E79")

    for col_idx, (titre, larg) in enumerate(zip(entetes, largeurs), start=1):
        cell = ws_liste.cell(row=4, column=col_idx, value=titre)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_liste.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_liste.row_dimensions[4].height = 20

    COULEURS_RESULTAT = {
        "conforme":     "C6EFCE",
        "non_conforme": "FFEB9C",
        "rebut":        "FFC7CE",
    }

    for row_idx, r in enumerate(retouches, start=5):
        res = r.get("resultat", "")
        couleur = COULEURS_RESULTAT.get(res, "FFFFFF")
        fill_res = PatternFill("solid", fgColor=couleur)
        valeurs = [
            r.get("id", ""),
            r.get("date", ""),
            r.get("operateur", ""),
            r.get("poste_travail", ""),
            r.get("defaut_constate", ""),
            r.get("operation", ""),
            RESULTATS.get(res, res),
            r.get("commentaire", ""),
        ]
        for col_idx, valeur in enumerate(valeurs, start=1):
            cell = ws_liste.cell(row=row_idx, column=col_idx, value=valeur)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 7:
                cell.fill = fill_res
                cell.font = Font(bold=True)
        ws_liste.row_dimensions[row_idx].height = 30

    ws_liste.freeze_panes = "A5"

    # ── Feuille 2 : Statistiques ──────────────────────────────────────────────
    ws_stats = wb.create_sheet("Statistiques")
    ws_stats["A1"] = f"STATISTIQUES RETOUCHES — {ref_article}"
    ws_stats["A1"].font = Font(bold=True, size=13)
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 20

    FILL_STAT = PatternFill("solid", fgColor="2E75B6")
    FILL_ALT  = PatternFill("solid", fgColor="D6E4F7")

    def _stat_row(ws, row, label, valeur, bold=False, fill=None):
        for col, val in ((1, label), (2, valeur)):
            c = ws.cell(row=row, column=col, value=val)
            if bold:
                c.font = Font(bold=True)
            if fill:
                c.fill = fill
            c.alignment = Alignment(horizontal="left")

    stats_lignes = [
        ("Total retouches",    stats["total"],                   True, FILL_STAT),
        ("Conformes",          stats["conformes"],                False, None),
        ("Non conformes",      stats["non_conformes"],            False, None),
        ("Rebuts",             stats["rebuts"],                   False, None),
        ("Taux de conformite", f"{stats['taux_conformite']:.1%}", True, FILL_ALT),
    ]
    for i, (label, valeur, bold, fill) in enumerate(stats_lignes, start=3):
        _stat_row(ws_stats, i, label, valeur, bold=bold, fill=fill)
        if bold and fill is FILL_STAT:
            ws_stats.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
            ws_stats.cell(row=i, column=2).font = Font(bold=True, color="FFFFFF")

    row = 10
    ws_stats.cell(row=row, column=1, value="Retouches par operateur").font = Font(bold=True)
    row += 1
    for op, count in sorted(stats["par_operateur"].items(), key=lambda x: -x[1]):
        ws_stats.cell(row=row, column=1, value=op)
        ws_stats.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws_stats.cell(row=row, column=1, value="Retouches par poste").font = Font(bold=True)
    row += 1
    for poste, count in sorted(stats["par_poste"].items(), key=lambda x: -x[1]):
        ws_stats.cell(row=row, column=1, value=poste)
        ws_stats.cell(row=row, column=2, value=count)
        row += 1

    wb.save(chemin_sortie)
    logger.info("Export retouches : %s (%d fiches)", chemin_sortie, len(retouches))
    return chemin_sortie
