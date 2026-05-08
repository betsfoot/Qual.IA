"""
Crée un fichier Excel de test avec des noms de colonnes volontairement différents
du standard Qual.IA — pour tester la détection automatique à l'import.
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "exports" / "TEST_Import_Client_Fictif.xlsx"
OUTPUT.parent.mkdir(exist_ok=True)

wb = openpyxl.Workbook()

# ─── Styles ──────────────────────────────────────────────────────────────────
BLEU     = "1A6BFF"
BLANC    = "FFFFFF"
GRIS     = "F2F4F8"
VERT     = "00C27A"
ORANGE   = "FF6B2B"

def style_header(cell, bg=BLEU):
    cell.font = Font(bold=True, color=BLANC, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

def style_data(cell, bg=BLANC):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="EEEEEE"),
        right=Side(style="thin", color="EEEEEE"),
    )

def ecrire_tableau(ws, headers, rows, start_row=2, bg_header=BLEU):
    ws.row_dimensions[start_row].height = 35
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        style_header(cell, bg_header)
    for r_idx, row in enumerate(rows, start_row + 1):
        bg = BLANC if r_idx % 2 == 0 else GRIS
        ws.row_dimensions[r_idx].height = 40
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            style_data(cell, bg)

# ═════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — "Analyse Produit" (= AMDEC Produit, noms volontairement différents)
# Colonnes : N° Ligne / Catégorie / Function / KCC / Failure Mode / Impact /
#            Severity / Root Cause / Frequency / Detection Method / Detection /
#            Criticality / Corrective Plan / Owner / Target Date
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Analyse Produit"

ws1["A1"] = "FMEA PRODUIT — Glace Saphir Ø28mm Chromée AR Simple"
ws1["A1"].font = Font(bold=True, size=13, color="0D1B3E")

headers_ap = [
    "N° Ligne", "Catégorie", "Function", "KCC",
    "Failure Mode", "Impact Client", "Severity",
    "Root Cause", "Frequency", "Detection Method", "Detection",
    "Criticality", "Corrective Plan", "Owner", "Target Date"
]

rows_ap = [
    ["01", "geometrique",  "Etancheite montre",          "Planeite",
     "Gauchissement glace",             "Infiltration humidite",    8,
     "Contrainte montage / defaut matiere",     4, "Controle planeite marbre",    3,
     "Action corrective",  "Ajuster couple montage",           "BT Qualite",  "J+15"],

    ["02", "optique",      "Transmission lumineuse",      "Taux transmission",
     "Rayure couche chrome",            "Refus client / retouche",  7,
     "Manipulation incorrecte",                5, "Inspection eclairage rasant", 2,
     "Action corrective",  "Protocole gants + supports Shore A",  "Resp Methodes","J+7"],

    ["03", "esthetique",   "Aspect visuel cadran",        "Aspect surface",
     "Piqure / eclat bord",             "Non-conformite visuelle",  6,
     "Choc manutention / emballage",           3, "Controle 100% loupe x10",    3,
     "Surveillance",       "Emballage individuel ESD",          "Logistique",  "J+30"],

    ["04", "optique",      "Antireflet efficace",         "Reflexion residuelle",
     "Delaminage couche AR",            "Perte transmission AR",    8,
     "Humidite salle blanche hors spec",       2, "Spectrophotometre 100%",     2,
     "Action corrective",  "Controle HR salle blanche < 45%",   "Resp Process","J+10"],

    ["05", "geometrique",  "Montage dans la lunette",     "Diametre",
     "Diametre hors tolerance",         "Impossibilite de montage", 9,
     "Derive outil rectification",             3, "Mesure pied coulisse 100%",  1,
     "Action corrective",  "Etalonnage journalier outil",       "Regleur",     "J+1"],

    ["06", "tracabilite",  "Suivi de production",         "Numero de lot",
     "Absence marquage tracabilite",    "Intraçabilite lot",        5,
     "Oubli operateur / bug ERP",              2, "Verification etiquette 100%",1,
     "Acceptable",         "Check-list avant liberation",       "BT Qualite",  "J+15"],

    ["07", "esthetique",   "Proprete surface",            "Absence particule",
     "Contamination particulaire",      "Point noir / fibre visible",6,
     "Nettoyage insuffisant avant depot",      4, "Inspection salle blanche",   2,
     "Action corrective",  "Double nettoyage US + ionisation",  "Resp Process","J+7"],
]

ecrire_tableau(ws1, headers_ap, rows_ap, start_row=2, bg_header=BLEU)

# Largeurs
largeurs_ap = [8, 12, 22, 20, 28, 25, 8, 32, 10, 25, 10, 16, 35, 14, 12]
for i, w in enumerate(largeurs_ap, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ═════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — "FMEA Process" (= AMDEC Process, noms encore différents)
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("FMEA Process")

ws2["A1"] = "FMEA PROCESS — Glace Saphir Ø28mm Chromée AR Simple"
ws2["A1"].font = Font(bold=True, size=13, color="0D1B3E")

headers_apr = [
    "Step #", "Process Step", "Station / Equipment",
    "Failure Mode", "Product Effect", "Sev",
    "Process Cause", "Occ", "Process Control", "Det",
    "RPN Class", "Action Plan", "Responsible", "Due Date",
    "Key Parameter", "Target Value"
]

rows_apr = [
    ["01", "Reception matiere",        "Zone reception",
     "Dimension hors plan",            "Rejet entree",              8,
     "Erreur fournisseur",                      3, "Controle dim entrant 100%",  2,
     "Action corrective",  "Retour fournisseur immediat",       "Appro",       "J+1",
     "Diametre entrant",   "28.00 +0/-0.05mm"],

    ["02", "Nettoyage ultrason",        "Cuve US 40kHz",
     "Residus apres nettoyage",        "Defaut adhesion chrome",     7,
     "Bain US sature / temperature basse",      4, "Controle visuel post-US",    3,
     "Action corrective",  "Renouveler bain US toutes 8h",      "Operateur",   "Continu",
     "Temperature bain",   "55°C ± 5°C"],

    ["03", "Activation plasma",         "Plasma cleaner Ar",
     "Activation insuffisante",        "Mauvaise adhesion couche",   8,
     "Pression Ar hors spec / electrode usee",  3, "Controle adhesion cross-cut",2,
     "Action corrective",  "Verification pression + electrode",  "Resp Process","Hebdo",
     "Puissance plasma",   "150W ± 10W"],

    ["04", "Depot Chrome sputtering",   "Bati sputtering Cr",
     "Epaisseur chrome non conforme",  "Reflexion hors spec",        8,
     "Derive puissance cathode / pression Ar",  3, "Mesure profilometre lot",    2,
     "Action corrective",  "Etalonnage profilometre + SPC",     "Resp Process","Hebdo",
     "Epaisseur Cr cible", "120nm ± 15nm"],

    ["05", "Antireflet PVD face 1",     "Evaporateur thermique",
     "Mauvaise transmission AR",       "Reflexion residuelle elevee", 8,
     "Derive temperature evaporation", 2,           "Spectropho 100% post-depot", 1,
     "Action corrective",  "Recalibrage courbe evaporation",    "Resp PVD",    "J+3",
     "Transmission AR F1", ">= 95%"],

    ["06", "Controle final 100%",       "Table controle eclairage rasant",
     "Defaut non detecte",             "Reclamation client",         9,
     "Fatigue operateur / eclairage insuffisant",2, "Controle visuel 100% + spectro",1,
     "Action corrective",  "Rotation poste 2h + etalon eclairage","BT Qualite","Continu",
     "Transmission finale",">=92% / Reflexion <=0.8%"],

    ["07", "Conditionnement",           "Poste emballage",
     "Rayure au conditionnement",      "Rebut post-liberation",      7,
     "Film protecteur absent / mal pose",       3, "Controle aspect avant emballage",2,
     "Surveillance",       "Gabarit pose film + check-list",    "Logistique",  "J+30",
     "Film protecteur",    "PE 50 microns face AR"],
]

ecrire_tableau(ws2, headers_apr, rows_apr, start_row=2, bg_header="FF6B2B")

largeurs_apr = [8, 22, 22, 28, 25, 6, 35, 6, 28, 6, 16, 35, 14, 10, 20, 18]
for i, w in enumerate(largeurs_apr, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ═════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — "Plan de fabrication" (= Gamme, noms encore différents)
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Plan de fabrication")

ws3["A1"] = "PLAN DE FABRICATION — Glace Saphir Ø28mm Chromée AR Simple"
ws3["A1"].font = Font(bold=True, size=13, color="0D1B3E")

headers_g = [
    "Op", "Libelle operation", "Instructions detaillees",
    "Machine / Poste", "Outillage", "Variable 1", "Spec 1",
    "Variable 2", "Spec 2", "Variable 3",
    "Tps (min)", "Point de controle", "Instrument", "Sampling",
    "Critere GO/NOGO", "Ref DIT", "Code ERP"
]

rows_g = [
    ["10", "Reception et controle entrant",
     "Verification certificat fournisseur, controle dimensionnel 100% au pied a coulisse",
     "Zone reception",        "Pied a coulisse 0.001mm + micrometre",
     "Diametre",              "28.00 +0/-0.05mm",
     "Epaisseur",             "2.50 ±0.03mm",
     "Aspect",                15, "Cotes + certificat + aspect",
     "Pied a coulisse + visuel",    "100%",  "Conforme plan + certif",       "DIT-010", "OP-10"],

    ["20", "Marquage tracabilite",
     "Attribution numero lot, impression etiquette code-barres, saisie ERP",
     "Bureau reception",      "Imprimante etiquettes Zebra",
     "Format etiquette",      "REF+LOT+DATE",
     "Lisibilite scan",       "100% scan OK",
     "—",                     5, "Tracabilite lot",
     "Lecteur code-barres",         "100%",  "Scan valide + saisie ERP OK",  "DIT-020", "OP-20"],

    ["30", "Nettoyage ultrason",
     "3 bains successifs : detergent neutre 55°C / eau deionisee / isopropanol. Sechage air chaud.",
     "Cuve ultrason 40kHz",   "Panier inox anti-rayure",
     "Temperature bain 1",    "55°C ±5°C",
     "Duree cycle",           "15 min",
     "Conductivite eau DI",   20, "Proprete surface",
     "Inspection visuelle x10",     "100%",  "Aucun residu / particule",     "DIT-030", "OP-30"],

    ["40", "Inspection pre-depot",
     "Controle visuel en salle blanche avant entree PVD. Rejection de tout defaut surface.",
     "Sas salle blanche",     "Microscope x10/x50 + eclairage rasant",
     "Aspect surface",        "Zero defaut",
     "Proprete",              "ISO classe 6",
     "—",                     10,"Aspect + proprete",
     "Microscope x10",              "100%",  "Zero rayure / particule",      "DIT-040", "OP-40"],

    ["50", "Depot chrome par sputtering",
     "Activation plasma Ar 150W / 5min. Depot Cr par sputtering cathodique DC.",
     "Bati sputtering Cr",    "Porte-pieces planetaire",
     "Epaisseur Cr",          "120nm ±15nm",
     "Puissance cathode",     "300W ±10W",
     "Pression Ar",           35, "Epaisseur + adhesion",
     "Profilometre Dektak",         "1/lot",  "120nm ±15 + cross-cut OK",    "DIT-050", "OP-50"],

    ["70", "Antireflet face 1 PVD",
     "Depot multicouche SiO2/TiO2 par evaporation thermique sous vide.",
     "Evaporateur PVD",       "Masque face 1 plan",
     "Couche SiO2",           "Lambda/4",
     "Transmission F1",       ">=95%",
     "Vide résiduel",         40, "Transmission + aspect",
     "Spectrophotometre",           "100%",  "T>=95% / aspect OK",           "DIT-070", "OP-70"],

    ["100","Controle final 100%",
     "Controle 100% tous criteres : dimensions, aspect, transmission, reflexion avant liberation.",
     "Table controle",        "Microscope x50 + spectrophotometre",
     "Transmission totale",   ">=92%",
     "Reflexion residuelle",  "<=0.8%",
     "Aspect",                25, "Tous criteres",
     "Spectropho + visuel rasant",  "100%",  "Tous criteres conformes",      "DIT-100", "OP-100"],

    ["110","Liberation lot",
     "Verification dossier complet, signature BT, enregistrement ERP, edition certificat conformite.",
     "Bureau qualite",        "Check-list liberation",
     "Completude dossier",    "100%",
     "Signature BT",          "Obligatoire",
     "—",                     15, "Dossier + signature",
     "Check-list + ERP",            "100%",  "Dossier complet + signe",      "DIT-110", "OP-110"],

    ["120","Conditionnement individuel",
     "Pose film PE 50 microns face AR, emballage sachet ESD individuel, etiquetage, mise en boite.",
     "Poste emballage",       "Film PE + sachets ESD + boite rigide",
     "Film protecteur",       "PE 50 microns",
     "Etiquette sachet",      "REF+LOT+QTE",
     "—",                     20, "Aspect emballage + etiquette",
     "Visuel",                      "100%",  "Film OK + etiquette lisible",  "DIT-120", "OP-120"],
]

ecrire_tableau(ws3, headers_g, rows_g, start_row=2, bg_header="00C27A")

largeurs_g = [6, 25, 45, 22, 28, 18, 18, 18, 18, 18, 8, 28, 25, 10, 28, 10, 10]
for i, w in enumerate(largeurs_g, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ─── Sauvegarde ──────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Fichier cree : {OUTPUT}")
print()
print("Onglets crees :")
print("  - 'Analyse Produit'    (AMDEC Produit  — colonnes en anglais/different)")
print("  - 'FMEA Process'       (AMDEC Process  — colonnes en anglais)")
print("  - 'Plan de fabrication'(Gamme           — colonnes en francais different)")
