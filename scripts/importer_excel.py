"""
Qual.IA — Import Excel client vers la base de références JSON.

Usage :
    python scripts/importer_excel.py

Le script guide pas à pas :
  1. Choisir la catégorie cible
  2. Choisir le fichier Excel client
  3. Mapper les onglets (AMDEC Produit / AMDEC Process / Gamme)
  4. Mapper les colonnes (détection automatique + correction manuelle)
  5. Renseigner les métadonnées (désignation, dimensions, traitements)
  6. Sauvegarder dans la base

Compatible avec n'importe quel Excel client, pas seulement le template Qual.IA.
"""

import json
import sys
import os
from pathlib import Path
from datetime import date

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
import pandas as pd

from backend.category_manager import lister_categories, charger_categorie
from backend.reference_saver import proposer_code_reference, enregistrer_reference, reference_existe

# ─── Colonnes cibles pour chaque document ────────────────────────────────────

COLONNES_AMDEC_PRODUIT = [
    "no", "famille", "fonction_exigence", "caracteristique_critique",
    "mode_defaillance", "effets_defaillance", "G", "causes_defaillance",
    "O", "controles_existants", "D", "classe_criticite",
    "actions_correctives", "responsable", "delai",
]

COLONNES_AMDEC_PROCESS = [
    "no", "operation_process", "etape_poste", "mode_defaillance",
    "effets_produit", "G", "causes_process", "O", "controles_process",
    "D", "classe_criticite", "actions_correctives", "responsable",
    "delai", "parametre_cle", "valeur_cible",
]

COLONNES_GAMME = [
    "no_op", "designation", "description_detaillee", "poste_machine",
    "outillage_fixture", "parametre_1", "valeur_1", "parametre_2",
    "valeur_2", "parametre_3", "temps_min", "point_controle",
    "moyen_controle", "frequence", "critere_acceptation",
    "ref_dit", "ref_infor",
]

# Synonymes courants pour la détection automatique
SYNONYMES = {
    # AMDEC Produit & Process communs
    "no":                      ["n°", "no", "num", "numéro", "#", "id", "ligne"],
    "famille":                 ["famille", "family", "categorie", "type"],
    "fonction_exigence":       ["fonction", "exigence", "function", "requirement", "fonction / exigence"],
    "caracteristique_critique":["caracteristique", "critique", "cct", "kpc", "caract"],
    "mode_defaillance":        ["mode", "défaillance", "failure mode", "fm", "mode de défaillance"],
    "effets_defaillance":      ["effet", "effects", "impact", "consequence", "effets"],
    "effets_produit":          ["effet", "effects", "impact", "effets produit"],
    "G":                       ["g", "gravité", "severity", "sev", "s"],
    "causes_defaillance":      ["cause", "causes", "root cause", "origine"],
    "causes_process":          ["cause", "causes process", "origine process"],
    "O":                       ["o", "occurrence", "freq", "fréquence occurrence", "probabilité"],
    "controles_existants":     ["controle", "control", "detection", "maîtrise", "contrôles existants"],
    "controles_process":       ["controle process", "maîtrise process", "control process"],
    "D":                       ["d", "détection", "detection", "det"],
    "classe_criticite":        ["criticite", "criticité", "classe", "criticality", "ipr", "npn"],
    "actions_correctives":     ["action", "corrective", "plan action", "actions correctives"],
    "responsable":             ["responsable", "resp", "owner", "pilote"],
    "delai":                   ["delai", "délai", "deadline", "echeance", "date"],
    # AMDEC Process spécifique
    "operation_process":       ["operation", "opération", "process", "étape", "step"],
    "etape_poste":             ["poste", "machine", "station", "etape", "poste / machine"],
    "parametre_cle":           ["parametre", "paramètre", "param", "kpc", "variable"],
    "valeur_cible":            ["valeur", "cible", "target", "spec", "valeur cible"],
    # Gamme
    "no_op":                   ["n° op", "no op", "op", "opération", "num op", "no_op"],
    "designation":             ["designation", "désignation", "nom operation", "libelle"],
    "description_detaillee":   ["description", "detail", "détail", "contenu", "instructions"],
    "poste_machine":           ["poste", "machine", "équipement", "equipment", "station"],
    "outillage_fixture":       ["outillage", "outil", "fixture", "jig", "tool"],
    "parametre_1":             ["parametre 1", "param 1", "paramètre 1", "p1"],
    "valeur_1":                ["valeur 1", "val 1", "v1"],
    "parametre_2":             ["parametre 2", "param 2", "paramètre 2", "p2"],
    "valeur_2":                ["valeur 2", "val 2", "v2"],
    "parametre_3":             ["parametre 3", "param 3", "paramètre 3", "p3"],
    "temps_min":               ["temps", "time", "duree", "durée", "minutes", "tps", "temps (min)"],
    "point_controle":          ["point controle", "point de contrôle", "checkpoint", "controle"],
    "moyen_controle":          ["moyen", "instrument", "outil controle", "moyen de contrôle"],
    "frequence":               ["frequence", "fréquence", "freq", "sampling"],
    "critere_acceptation":     ["critere", "critère", "acceptation", "acceptance", "spec"],
    "ref_dit":                 ["ref dit", "dit", "document", "ref doc", "réf. dit"],
    "ref_infor":               ["ref infor", "infor", "erp", "ref erp", "réf. infor"],
}


# ─── Utilitaires ─────────────────────────────────────────────────────────────

def _normaliser(texte: str) -> str:
    import unicodedata
    if not isinstance(texte, str):
        return ""
    t = texte.lower().strip()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t


def _detecter_colonne(col_excel: str, cible: str) -> float:
    """Score 0–1 : à quel point col_excel ressemble à la colonne cible."""
    norm_excel = _normaliser(col_excel)
    synonymes = [_normaliser(s) for s in SYNONYMES.get(cible, [cible])]
    for syn in synonymes:
        if norm_excel == syn:
            return 1.0
        if syn in norm_excel or norm_excel in syn:
            return 0.7
    return 0.0


def _mapper_colonnes_auto(colonnes_excel: list, colonnes_cibles: list) -> dict:
    """
    Retourne un dict {cible: colonne_excel} pour les correspondances automatiques.
    Seuil : score >= 0.7
    """
    mapping = {}
    for cible in colonnes_cibles:
        meilleur_score = 0
        meilleure_col = None
        for col in colonnes_excel:
            score = _detecter_colonne(str(col), cible)
            if score > meilleur_score:
                meilleur_score = score
                meilleure_col = col
        if meilleur_score >= 0.7:
            mapping[cible] = meilleure_col
    return mapping


def _lire_sheet(wb: openpyxl.Workbook, nom_sheet: str, ligne_header: int = 1) -> pd.DataFrame | None:
    """Lit un onglet et retourne un DataFrame. ligne_header = numéro de ligne (1-based) des en-têtes."""
    if nom_sheet not in wb.sheetnames:
        return None
    ws = wb[nom_sheet]
    data = list(ws.values)
    if not data or len(data) <= ligne_header:
        return None
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(data[ligne_header - 1])]
    rows = data[ligne_header:]
    df = pd.DataFrame(rows, columns=headers)
    # Supprimer lignes entièrement vides
    df = df.dropna(how="all")
    return df


def _demander_sheet(wb: openpyxl.Workbook, doc_type: str) -> str | None:
    """Demande à l'utilisateur quel onglet correspond au document."""
    sheets = wb.sheetnames
    print(f"\nOnglets disponibles dans le fichier :")
    for i, s in enumerate(sheets):
        print(f"  [{i+1}] {s}")
    print(f"  [0] Aucun — ignorer {doc_type}")
    choix = input(f"Quel onglet contient l'{doc_type} ? > ").strip()
    if choix == "0" or choix == "":
        return None
    try:
        idx = int(choix) - 1
        if 0 <= idx < len(sheets):
            return sheets[idx]
    except ValueError:
        # L'utilisateur a tapé le nom directement
        if choix in sheets:
            return choix
    print("Choix invalide — onglet ignoré.")
    return None


def _corriger_mapping(colonnes_excel: list, colonnes_cibles: list, mapping_auto: dict, doc_type: str) -> dict:
    """Affiche le mapping auto et permet de corriger."""
    print(f"\n--- Mapping colonnes pour {doc_type} ---")
    print(f"Colonnes détectées dans Excel : {colonnes_excel}")
    print(f"\nMapping automatique :")

    non_mappes = [c for c in colonnes_cibles if c not in mapping_auto]
    for cible in colonnes_cibles:
        col = mapping_auto.get(cible)
        statut = f"-> {col}" if col else "!! NON DÉTECTÉ"
        print(f"  {cible:35s} {statut}")

    if non_mappes:
        print(f"\n{len(non_mappes)} colonne(s) non détectée(s) : {non_mappes}")
        corriger = input("Voulez-vous corriger le mapping manuellement ? (o/N) > ").strip().lower()
        if corriger == "o":
            for cible in non_mappes:
                print(f"\nColonnes Excel disponibles :")
                for i, c in enumerate(colonnes_excel):
                    print(f"  [{i+1}] {c}")
                print(f"  [0] Laisser vide pour '{cible}'")
                choix = input(f"Quelle colonne pour '{cible}' ? > ").strip()
                if choix == "0" or choix == "":
                    continue
                try:
                    idx = int(choix) - 1
                    if 0 <= idx < len(colonnes_excel):
                        mapping_auto[cible] = colonnes_excel[idx]
                except ValueError:
                    if choix in colonnes_excel:
                        mapping_auto[cible] = choix
    return mapping_auto


def _appliquer_mapping(df: pd.DataFrame, mapping: dict, colonnes_cibles: list) -> list[dict]:
    """Applique le mapping et retourne une liste de dicts."""
    lignes = []
    for _, row in df.iterrows():
        ligne = {}
        for cible in colonnes_cibles:
            col_excel = mapping.get(cible)
            if col_excel and col_excel in df.columns:
                val = row[col_excel]
                if pd.isna(val):
                    val = ""
                # Convertir G, O, D en int
                if cible in ("G", "O", "D"):
                    try:
                        val = int(float(val)) if val != "" else ""
                    except (ValueError, TypeError):
                        val = ""
                elif cible in ("temps_min",):
                    try:
                        val = float(val) if val != "" else 0
                    except (ValueError, TypeError):
                        val = 0
                else:
                    val = str(val).strip() if val != "" else ""
            else:
                val = ""
            ligne[cible] = val
        # Ignorer les lignes complètement vides
        if any(v for v in ligne.values() if str(v).strip()):
            lignes.append(ligne)
    return lignes


def _saisir_metadata(categorie) -> dict:
    """Demande les métadonnées de la référence à importer."""
    print("\n=== Métadonnées de la référence ===")

    designation = input("Désignation complète (ex: Glace ronde saphir Ø34mm CN+AR) : ").strip()

    vocab = categorie.vocabulaire()
    types = vocab.get("types_produit", {})
    matieres = vocab.get("matieres", {})
    traitements_dispo = vocab.get("traitements", {})

    # Type produit
    print(f"\nTypes disponibles pour {categorie.nom} :")
    types_list = list(types.keys())
    for i, (k, v) in enumerate(types.items()):
        print(f"  [{i+1}] {k} — {v}")
    print(f"  [0] Saisir manuellement")
    choix = input("Type produit > ").strip()
    try:
        idx = int(choix) - 1
        type_produit = types_list[idx] if 0 <= idx < len(types_list) else choix
    except ValueError:
        type_produit = choix

    # Matière
    print(f"\nMatières disponibles :")
    matieres_list = list(matieres.keys())
    for i, (k, v) in enumerate(matieres.items()):
        print(f"  [{i+1}] {k} — {v}")
    print(f"  [0] Saisir manuellement")
    choix = input("Matière > ").strip()
    try:
        idx = int(choix) - 1
        matiere = matieres_list[idx] if 0 <= idx < len(matieres_list) else choix
    except ValueError:
        matiere = choix

    # Traitements
    print(f"\nTraitements disponibles (séparer par virgule, ex: 1,3) :")
    traitements_list = list(traitements_dispo.keys())
    for i, (k, v) in enumerate(traitements_dispo.items()):
        print(f"  [{i+1}] {k} — {v}")
    choix = input("Traitements (numéros ou noms) > ").strip()
    traitements = []
    for c in choix.split(","):
        c = c.strip()
        try:
            idx = int(c) - 1
            if 0 <= idx < len(traitements_list):
                traitements.append(traitements_list[idx])
        except ValueError:
            if c:
                traitements.append(c)

    # Dimensions
    print("\nDimensions :")
    champs = categorie.champs_brief()
    try:
        diametre = float(input(f"  Diamètre principal (mm) [{champs.get('diametre_min',1)}–{champs.get('diametre_max',200)}] : "))
    except ValueError:
        diametre = 0.0
    try:
        tolerance_d = float(input(f"  Tolérance diamètre (mm) [défaut {champs.get('tolerance_diametre_default',0.05)}] : ") or champs.get("tolerance_diametre_default", 0.05))
    except ValueError:
        tolerance_d = float(champs.get("tolerance_diametre_default", 0.05))
    try:
        epaisseur = float(input(f"  Épaisseur (mm) [0 si inconnue] : ") or 0)
    except ValueError:
        epaisseur = 0.0
    try:
        tolerance_e = float(input(f"  Tolérance épaisseur (mm) [défaut {champs.get('tolerance_epaisseur_default',0.03)}] : ") or champs.get("tolerance_epaisseur_default", 0.03))
    except ValueError:
        tolerance_e = float(champs.get("tolerance_epaisseur_default", 0.03))

    # Statut
    statut = input("\nStatut [valide/brouillon] (défaut: valide) : ").strip() or "valide"
    approuve_par = input("Approuvé par (optionnel) : ").strip() or None

    brief = {
        "type_produit": type_produit,
        "matiere": matiere,
        "traitements": traitements,
        "dimensions": {
            "diametre_mm": diametre,
            "tolerance_diametre_mm": tolerance_d,
            "epaisseur_mm": epaisseur if epaisseur > 0 else None,
            "tolerance_epaisseur_mm": tolerance_e,
        },
        "designation_client": designation,
        "exigences_speciales": "",
    }
    return brief, statut, approuve_par, designation


def _detecter_ligne_header(ws) -> int:
    """Détecte la ligne d'en-tête en cherchant la ligne avec le plus de cellules non vides."""
    scores = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        non_vides = sum(1 for c in row if c is not None and str(c).strip())
        scores[i] = non_vides
    return max(scores, key=scores.get) if scores else 1


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Qual.IA — Import Excel client")
    print("=" * 60)

    # 1. Choisir la catégorie
    categories = lister_categories()
    print("\nCatégories disponibles :")
    for i, cat in enumerate(categories):
        print(f"  [{i+1}] {cat.icone} {cat.nom}")
    choix = input("Catégorie cible > ").strip()
    try:
        idx = int(choix) - 1
        categorie = categories[idx]
    except (ValueError, IndexError):
        print("Choix invalide.")
        sys.exit(1)
    print(f"Catégorie sélectionnée : {categorie.icone} {categorie.nom}")

    # 2. Choisir le fichier Excel
    chemin_excel = input("\nChemin du fichier Excel client (glisser-déposer possible) : ").strip().strip('"')
    chemin_excel = Path(chemin_excel)
    if not chemin_excel.exists():
        print(f"Fichier introuvable : {chemin_excel}")
        sys.exit(1)

    print(f"Ouverture de : {chemin_excel.name}")
    wb = openpyxl.load_workbook(chemin_excel, data_only=True)
    print(f"Onglets trouvés : {wb.sheetnames}")

    # 3. Mapper les onglets
    print("\n--- Identification des onglets ---")
    print("Indiquer quel onglet contient chaque document :")

    # Tentative de détection automatique par nom
    def _auto_sheet(mot_cle):
        for s in wb.sheetnames:
            if mot_cle.lower() in s.lower():
                return s
        return None

    sheet_ap  = _auto_sheet("produit") or _auto_sheet("AMDEC P") or _auto_sheet("AP")
    sheet_apr = _auto_sheet("process") or _auto_sheet("AMDEC PR") or _auto_sheet("APR")
    sheet_g   = _auto_sheet("gamme") or _auto_sheet("Gamme") or _auto_sheet("G")

    for doc, sheet_auto in [
        ("AMDEC Produit", sheet_ap),
        ("AMDEC Process", sheet_apr),
        ("Gamme", sheet_g),
    ]:
        if sheet_auto:
            conf = input(f"  {doc} → onglet '{sheet_auto}' détecté automatiquement. Confirmer ? (O/n) > ").strip().lower()
            if conf == "n":
                sheet_auto = _demander_sheet(wb, doc)
        else:
            print(f"  {doc} non détecté automatiquement.")
            sheet_auto = _demander_sheet(wb, doc)

        if doc == "AMDEC Produit":
            sheet_ap = sheet_auto
        elif doc == "AMDEC Process":
            sheet_apr = sheet_auto
        else:
            sheet_g = sheet_auto

    # 4. Lire et mapper chaque document
    resultats = {}

    for doc_type, sheet_name, colonnes_cibles, cle_json in [
        ("AMDEC Produit",  sheet_ap,  COLONNES_AMDEC_PRODUIT,  "amdec_produit"),
        ("AMDEC Process",  sheet_apr, COLONNES_AMDEC_PROCESS,  "amdec_process"),
        ("Gamme",          sheet_g,   COLONNES_GAMME,          "gamme"),
    ]:
        if not sheet_name:
            print(f"\n[SKIP] {doc_type} — onglet non défini.")
            resultats[cle_json] = {"modes_defaillance": [] if "amdec" in cle_json else None, "operations": []}
            continue

        ws = wb[sheet_name]
        ligne_header = _detecter_ligne_header(ws)
        df = _lire_sheet(wb, sheet_name, ligne_header)

        if df is None or df.empty:
            print(f"\n[SKIP] {doc_type} — onglet vide ou illisible.")
            continue

        colonnes_excel = [c for c in df.columns if c and not str(c).startswith("col_")]
        mapping_auto = _mapper_colonnes_auto(colonnes_excel, colonnes_cibles)
        mapping = _corriger_mapping(colonnes_excel, colonnes_cibles, mapping_auto, doc_type)
        lignes = _appliquer_mapping(df, mapping, colonnes_cibles)

        print(f"\n[OK] {doc_type} : {len(lignes)} lignes importées depuis l'onglet '{sheet_name}'")
        resultats[cle_json] = lignes

    # 5. Métadonnées
    brief, statut, approuve_par, designation = _saisir_metadata(categorie)

    # 6. Code de référence
    code_propose = proposer_code_reference(categorie, brief)
    print(f"\nCode proposé : {code_propose}")
    code_ref = input(f"Code de référence (Entrée pour accepter '{code_propose}') : ").strip() or code_propose

    if reference_existe(categorie, code_ref):
        ecr = input(f"La référence '{code_ref}' existe déjà. Écraser ? (o/N) > ").strip().lower()
        overwrite = ecr == "o"
        if not overwrite:
            print("Import annulé.")
            sys.exit(0)
    else:
        overwrite = False

    # 7. Construire le dossier JSON
    dossier = {
        "amdec_produit": {
            "reference": code_ref,
            "document": "AMDEC Produit",
            "designation": designation,
            "modes_defaillance": resultats.get("amdec_produit", []),
            "confiance_globale": 1.0,
            "avertissements_generateur": ["Importé depuis Excel client"],
        },
        "amdec_process": {
            "reference": code_ref,
            "document": "AMDEC Process",
            "designation": designation,
            "modes_defaillance": resultats.get("amdec_process", []),
            "confiance_globale": 1.0,
            "avertissements_generateur": ["Importé depuis Excel client"],
        },
        "gamme": {
            "reference": code_ref,
            "document": "Gamme de Production",
            "designation": designation,
            "operations": resultats.get("gamme", []),
            "confiance_globale": 1.0,
            "avertissements_generateur": ["Importé depuis Excel client"],
        },
        "metadonnees_generation": {
            "reference_source": "import_excel",
            "score_similarite": 1.0,
            "mode_generation": "import",
        },
    }

    # 8. Enregistrer (workflow auto-initialisé en brouillon)
    ref_dir = enregistrer_reference(
        categorie=categorie,
        code=code_ref,
        brief=brief,
        dossier=dossier,
        acteur=approuve_par or "Import CLI",
        commentaire_creation=f"Import CLI — {code_ref}",
        overwrite=overwrite,
    )

    print(f"\n{'='*60}")
    print(f"  Import termine avec succes !")
    print(f"  Reference : {code_ref}")
    print(f"  Dossier   : {ref_dir}")
    nb_ap  = len(resultats.get("amdec_produit", []))
    nb_apr = len(resultats.get("amdec_process", []))
    nb_g   = len(resultats.get("gamme", []))
    print(f"  AMDEC Produit : {nb_ap} modes")
    print(f"  AMDEC Process : {nb_apr} modes")
    print(f"  Gamme         : {nb_g} operations")
    print(f"{'='*60}")
    print("\nLa reference est maintenant disponible dans Qual.IA.")


if __name__ == "__main__":
    main()
